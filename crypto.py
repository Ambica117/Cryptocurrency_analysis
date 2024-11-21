from pycoingecko import CoinGeckoAPI
from openpyxl import Workbook, load_workbook
import time

def fetch_top_cryptos():
    cg = CoinGeckoAPI()
    response = cg.get_coins_markets(vs_currency='usd', order='market_cap_desc', per_page=50, page=1)
    return response

def update_excel():
   
    crypto_data = fetch_top_cryptos()

    try:
        workbook = load_workbook('crypto_data.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Crypto Name'
        sheet['B1'] = 'Symbol'
        sheet['C1'] = 'Price (USD)'
        sheet['D1'] = 'Market Cap'
        sheet['E1'] = '24h Volume'
        sheet['F1'] = 'Price Change (24h, %)'

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

    for i, crypto in enumerate(crypto_data, start=2):
        sheet[f'A{i}'] = crypto['name']
        sheet[f'B{i}'] = crypto['symbol']
        sheet[f'C{i}'] = crypto['current_price']
        sheet[f'D{i}'] = crypto['market_cap']
        sheet[f'E{i}'] = crypto['total_volume']
        sheet[f'F{i}'] = crypto['price_change_percentage_24h']

    workbook.save('crypto_data.xlsx')

while True:
    update_excel()
    print("Excel updated with live data.")
    time.sleep(300)  
